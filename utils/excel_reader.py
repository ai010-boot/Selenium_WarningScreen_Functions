"""
Excel 数据读取模块
用于读取 Excel 格式的测试数据
"""
import openpyxl
from pathlib import Path
from utils.logger import Logger

class ExcelReader:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.sheet_names = []
        
        try:
            # 使用openpyxl加载工作簿
            self.workbook = openpyxl.load_workbook(file_path, read_only=True)
            self.sheet_names = self.workbook.sheetnames
        except Exception as e:
            raise Exception(f"加载Excel文件失败: {e}")
    
    def get_sheet_data(self, sheet_name):
        """获取指定工作表的数据"""
        if self.workbook is None:
            raise Exception("工作簿未加载成功")
            
        if sheet_name not in self.sheet_names:
            raise ValueError(f"工作表 '{sheet_name}' 不存在")
            
        sheet = self.workbook[sheet_name]
        data = []
        
        # 获取标题行
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        
        # 过滤掉空的标题
        if not any(headers):
            return data
        
        # 获取数据行
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # 跳过全空行
            if not any(row):
                continue
            row_data = {}
            for i, value in enumerate(row):
                if i < len(headers):  # 防止索引超出范围
                    row_data[headers[i]] = value
            data.append(row_data)
            
        return data
    
    def get_cell_value(self, sheet_name, row, col):
        """获取指定单元格的值"""
        if self.workbook is None:
            raise Exception("工作簿未加载成功")
            
        if sheet_name not in self.sheet_names:
            raise ValueError(f"工作表 '{sheet_name}' 不存在")
            
        sheet = self.workbook[sheet_name]
        cell = sheet.cell(row=row, column=col)
        return cell.value
    
    def close(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()


def read_excel_data(filepath, sheet_name=None):
    """
    便捷函数：读取 Excel 数据
    
    Args:
        filepath: Excel 文件路径
        sheet_name: 工作表名称
    
    Returns:
        数据列表
    """
    reader = ExcelReader(filepath)
    data = reader.get_sheet_data(sheet_name)
    reader.close()
    return data
